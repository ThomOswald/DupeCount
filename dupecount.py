import openpyxl

class DupeCount:
    rowOfUser = []

    newBook = openpyxl.Workbook()

    def __init__(self, origFile):
        # File format matters! Make sure empty cells are empty before starting or you're gonna have a bad time.
        self.origFile = origFile
        origBook = openpyxl.load_workbook(self.origFile)
        origSheet = origBook.active  # keep naming convention consistent

        # loop through i up to the last row

        self.addColumn(origSheet)
            # print(user)

        self.columnToRow()
        self.newBook.save("Counted "+self.origFile)
        print("Your file", "Counted "+self.origFile+" is ready, my guy")

    def addColumn(self, origSheet):
        for row in origSheet.iter_rows(origSheet.min_row, origSheet.max_row, origSheet.min_column, origSheet.max_column):

            user = [col.value for col in row]

            #     user.append(col.value)  # add column to everything

            user.append(1)  # add column with value of 1

            for _ in range(1, 19):
                if user in self.rowOfUser:
                    user[origSheet.max_column] += 1
            else:
                pass
            self.rowOfUser.append(user)
            # print(user)

    def columnToRow(self):
        newSheet = self.newBook.active
        for value in self.rowOfUser:
            newSheet.append(value)

# example: DupeCount('listofsomesort.xlsx')

DupeCount()
